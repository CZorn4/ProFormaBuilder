import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from typing import Dict

def apply_financial_table_style(ws):
    """
    Apply consistent styling to Excel worksheets for financial data
    """
    # Define border style
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Header fill color (light blue)
    header_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
    section_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    # Fonts
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    
    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            # Left align text in first column, right align numbers in other columns
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')
            
            # Apply number formatting
            if isinstance(cell.value, (int, float)) and cell.column > 1:
                if 'rate' in str(ws.cell(row=cell.row, column=1).value).lower():
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00'

def format_valuation_sheet(ws):
    """
    Special formatting for the valuation sheet
    """
    # Section headers styling
    section_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    section_font = Font(bold=True, size=12)
    
    # Find and format section headers
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in ["WACC Components", "Present Value Calculations", "Valuation Summary"]:
                cell.font = section_font
                cell.fill = section_fill
                ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=4)
                cell.alignment = Alignment(horizontal='center')
    
    # Format WACC components
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column == 2:
                if any(rate in str(ws.cell(row=cell.row, column=1).value).lower() 
                      for rate in ['rate', 'wacc', 'weight']):
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00'
    
    # Format Present Value table
    pv_start = None
    for row in ws.iter_rows():
        if row[0].value == "Present Value Calculations":
            pv_start = row[0].row + 2  # Skip header row
            break
    
    if pv_start:
        for row in ws.iter_rows(min_row=pv_start):
            if not any(cell.value for cell in row):  # Empty row
                break
            for cell in row:
                if cell.column > 1:  # Skip year column
                    if cell.column == 3:  # PV Factor
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0.00'
    
    # Add conclusion section
    last_row = ws.max_row
    ws.append([])  # Empty row
    last_row += 2
    
    ws.cell(row=last_row, column=1, value="Valuation Conclusions")
    ws.cell(row=last_row, column=1).font = section_font
    ws.cell(row=last_row, column=1).fill = section_fill
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
    
    # Add lines for conclusions
    for i in range(5):
        ws.append([""])
        for col in range(1, 5):
            cell = ws.cell(row=last_row + 1 + i, column=col)
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin') if i == 0 else None,
                bottom=Side(style='thin') if i == 4 else None
            )

def export_financial_analysis_to_excel(
    financial_data: Dict, 
    ratio_data: Dict, 
    projected_data: Dict, 
    fcf_data: Dict, 
    wacc_data: Dict, 
    valuation_results: Dict, 
    output_file: str = 'financial_analysis.xlsx'
):
    """
    Export comprehensive financial analysis to Excel with horizontal, year-by-year tables
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Income Statement Order
    income_statement_order = [
        'revenue',
        'cost_of_goods_sold',
        'interest_paid',
        'interest_earned',
        'depreciation',
        'profit_before_tax',
        'income_tax_expense',
        'profit_after_tax',
        'dividends'
    ]
    
    # Balance Sheet Order
    balance_sheet_order = [
        'cash_and_securities',
        'total_current_assets',
        'ppe_gross',
        'accumulated_depreciation',
        'ppe_net',
        'total_assets',
        'current_liabilities',
        'total_liabilities',
        'outstanding_stock',
        'StockHolders_Equity',
        'retained_earnings',
        'total_liabilities_and_SE'
    ]
    
    # Ratio Order
    ratio_order = [
        'revenue_growth_rate',
        'current_assets_to_revenue',
        'current_liabilities_to_revenue',
        'net_fixed_assets_to_revenue',
        'operating_costs_to_revenue',
        'depreciation_rate',
        'interest_rate',
        'interest_paid_rate',
        'tax_rate',
        'dividend_growth_rate'
    ]
    
    # Free Cash Flow Order
    fcf_order = [
        'profit_after_tax',
        'add_back_depreciation',
        'subtract_increase_in_current_assets',
        'add_back_increase_in_current_liabilities',
        'subtract_increase_in_fixed_assets',
        'add_back_after_tax_interest_on_debt',
        'subtract_after_tax_interest_on_cash',
        'free_cash_flow'
    ]
    
    # Historical Financial Statements Sheet
    ws_historical = wb.active
    ws_historical.title = "Historical Financials"
    
    # Collect years
    historical_years = sorted(financial_data['years'].keys())
    
    # Income Statement Section
    ws_historical.cell(row=1, column=1, value="Income Statement")
    for col, year in enumerate(historical_years, start=2):
        ws_historical.cell(row=1, column=1+col, value=str(year))
    
    # Write Income Statement Items and Values
    for row, item in enumerate(income_statement_order, start=2):
        ws_historical.cell(row=row, column=1, value=item.replace('_', ' ').title())
        for col, year in enumerate(historical_years, start=2):
            value = financial_data['years'][year]['income_statement'].get(item, {}).get('value', '')
            ws_historical.cell(row=row, column=1+col, value=value)
    
    # Balance Sheet Section
    balance_sheet_offset = len(income_statement_order) + 3
    
    ws_historical.cell(row=balance_sheet_offset, column=1, value="Balance Sheet")
    for col, year in enumerate(historical_years, start=2):
        ws_historical.cell(row=balance_sheet_offset, column=1+col, value=str(year))
    
    # Write Balance Sheet Items and Values
    for row, item in enumerate(balance_sheet_order, start=balance_sheet_offset+1):
        ws_historical.cell(row=row, column=1, value=item.replace('_', ' ').title())
        for col, year in enumerate(historical_years, start=2):
            value = financial_data['years'][year]['balance_sheet'].get(item, {}).get('value', '')
            ws_historical.cell(row=row, column=1+col, value=value)
    
    # Financial Ratios Sheet
    ws_ratios = wb.create_sheet("Financial Ratios")
    ratio_years = sorted(ratio_data['years'].keys())
    
    ws_ratios.cell(row=1, column=1, value="Financial Ratios")
    for col, year in enumerate(ratio_years, start=2):
        ws_ratios.cell(row=1, column=1+col, value=str(year))
    
    # Write Ratio Items and Values
    for row, item in enumerate(ratio_order, start=2):
        ws_ratios.cell(row=row, column=1, value=item.replace('_', ' ').title())
        for col, year in enumerate(ratio_years, start=2):
            value = ratio_data['years'][year]['ratios'].get(item, {}).get('value', '')
            ws_ratios.cell(row=row, column=1+col, value=value)
    
    # Free Cash Flow Sheet
    ws_fcf = wb.create_sheet("Free Cash Flow")
    fcf_years = sorted(fcf_data['years'].keys())
    
    ws_fcf.cell(row=1, column=1, value="Free Cash Flow")
    for col, year in enumerate(fcf_years, start=2):
        ws_fcf.cell(row=1, column=1+col, value=str(year))
    
    # Write FCF Items and Values
    for row, item in enumerate(fcf_order, start=2):
        ws_fcf.cell(row=row, column=1, value=item.replace('_', ' ').title())
        for col, year in enumerate(fcf_years, start=2):
            value = fcf_data['years'][year]['fcf_calculation'].get(item, {}).get('value', '')
            ws_fcf.cell(row=row, column=1+col, value=value)
    
    # Projected Financials Sheet
    ws_projected = wb.create_sheet("Projected Financials")
    projected_years = sorted(projected_data['years'].keys())
    
    # Income Statement Projections
    ws_projected.cell(row=1, column=1, value="Projected Income Statement")
    for col, year in enumerate(projected_years, start=2):
        ws_projected.cell(row=1, column=1+col, value=str(year))
    
    for row, item in enumerate(income_statement_order, start=2):
        ws_projected.cell(row=row, column=1, value=item.replace('_', ' ').title())
        for col, year in enumerate(projected_years, start=2):
            value = projected_data['years'][year]['income_statement'].get(item, {}).get('value', '')
            ws_projected.cell(row=row, column=1+col, value=value)
    
    # Balance Sheet Projections
    balance_proj_offset = len(income_statement_order) + 3
    ws_projected.cell(row=balance_proj_offset, column=1, value="Projected Balance Sheet")
    for col, year in enumerate(projected_years, start=2):
        ws_projected.cell(row=balance_proj_offset, column=1+col, value=str(year))
    
    for row, item in enumerate(balance_sheet_order, start=balance_proj_offset+1):
        ws_projected.cell(row=row, column=1, value=item.replace('_', ' ').title())
        for col, year in enumerate(projected_years, start=2):
            value = projected_data['years'][year]['balance_sheet'].get(item, {}).get('value', '')
            ws_projected.cell(row=row, column=1+col, value=value)
    
    # Valuation and WACC Analysis Sheet
    ws_valuation = wb.create_sheet("Valuation & WACC")
    
    # WACC Components
    row = 1
    ws_valuation.cell(row=row, column=1, value="WACC Components")
    row += 1
    
    for component, value in wacc_data.items():
        ws_valuation.cell(row=row, column=1, value=component.replace('_', ' ').title())
        ws_valuation.cell(row=row, column=2, value=value)
        row += 1
    
    # Present Value Calculations
    row += 2
    ws_valuation.cell(row=row, column=1, value="Present Value Calculations")
    row += 1
    
    headers = ['Year', 'FCF', 'PV Factor', 'Present Value']
    for col, header in enumerate(headers, start=1):
        ws_valuation.cell(row=row, column=col, value=header)
    row += 1
    
    for year, pv_data in sorted(valuation_results['present_values'].items()):
        fcf = pv_data['fcf']
        pv = pv_data['present_value']
        pv_factor = pv / fcf if fcf != 0 else 0
        
        ws_valuation.cell(row=row, column=1, value=year)
        ws_valuation.cell(row=row, column=2, value=fcf)
        ws_valuation.cell(row=row, column=3, value=pv_factor)
        ws_valuation.cell(row=row, column=4, value=pv)
        row += 1
    
    # Valuation Summary
    row += 2
    ws_valuation.cell(row=row, column=1, value="Valuation Summary")
    row += 1
    
    summary = valuation_results['valuation_summary']
    metrics = [
        ('WACC', summary['wacc']),
        ('Long Term Growth Rate', summary['long_term_growth_rate']),
        ('Present Value of FCF', summary['pv_fcf_total']),
        ('Terminal Value', summary['terminal_value']),
        ('PV of Terminal Value', summary['pv_terminal_value']),
        ('Enterprise Value', summary['enterprise_value']),
        ('Cash and Securities', summary.get('cash_and_securities', 0)),
        ('Total Liabilities', summary.get('total_liabilities', 0)),
        ('Equity Value', summary['equity_value']),
        ('Shares Outstanding', summary['shares_outstanding']),
        ('Equity Value per Share', summary.get('equity_value_per_share', 'N/A'))
    ]
    
    for label, value in metrics:
        ws_valuation.cell(row=row, column=1, value=label)
        ws_valuation.cell(row=row, column=2, value=value)
        row += 1
    
    # Style all worksheets
    for sheet in wb.worksheets:
        apply_financial_table_style(sheet)
        
        # Special formatting for valuation sheet
        if sheet.title == "Valuation & WACC":
            format_valuation_sheet(sheet)
        
        # Adjust column widths (fixed version)
        for column_cells in sheet.columns:
            max_length = 0
            column = None
            
            for cell in column_cells:
                try:
                    if cell.coordinate:  # Check if cell has coordinate
                        if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Skip merged cells
                            if column is None:
                                column = cell.column_letter  # Get column letter from first non-merged cell
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                except:
                    pass
            
            if column:  # Only adjust if we found a valid column
                adjusted_width = (max_length + 2) * 1.2  # Added 20% more space
                sheet.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Financial analysis exported to {output_file}")